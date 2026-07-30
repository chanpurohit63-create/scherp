import io
import json
from typing import List
from datetime import datetime

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from . import models, schemas

DAYS_OF_WEEK = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


def generate_timetable_pdf(entries: List[models.Timetable]) -> bytes:
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "School Timetable", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 8, f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}", ln=True, align="C")
    pdf.ln(5)

    pdf.set_font("Helvetica", "B", 8)
    col_widths = [18, 25, 30, 30, 30, 25, 25, 25, 25, 25]
    headers = ["Day", "Period", "Start", "End", "Subject", "Teacher", "Class", "Section", "Room", "Status"]
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 7, h, border=1, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for entry in entries:
        row_data = [
            DAYS_OF_WEEK.get(entry.day_of_week, str(entry.day_of_week)),
            str(entry.period),
            entry.start_time or "",
            entry.end_time or "",
            str(entry.subject_id),
            str(entry.teacher_id),
            str(entry.class_id),
            str(entry.section_id or "-"),
            str(entry.room_id or "-"),
            entry.status,
        ]
        for i, val in enumerate(row_data):
            pdf.cell(col_widths[i], 6, val, border=1, align="C")
        pdf.ln()

    return pdf.output(dest="S").encode("latin-1")


def generate_timetable_excel(entries: List[models.Timetable]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Day", "Period", "Start Time", "End Time", "Subject ID", "Teacher ID", "Class ID", "Section ID", "Room ID", "Status", "Remarks"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, entry in enumerate(entries, 2):
        row_data = [
            DAYS_OF_WEEK.get(entry.day_of_week, str(entry.day_of_week)),
            entry.period,
            entry.start_time or "",
            entry.end_time or "",
            entry.subject_id,
            entry.teacher_id,
            entry.class_id,
            entry.section_id or "",
            entry.room_id or "",
            entry.status,
            entry.remarks or "",
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_timetable_csv(entries: List[models.Timetable]) -> str:
    output = io.StringIO()
    output.write("Day,Period,Start Time,End Time,Subject ID,Teacher ID,Class ID,Section ID,Room ID,Status,Remarks\n")
    for entry in entries:
        output.write(
            f"{DAYS_OF_WEEK.get(entry.day_of_week, entry.day_of_week)},"
            f"{entry.period},"
            f"{entry.start_time or ''},"
            f"{entry.end_time or ''},"
            f"{entry.subject_id},"
            f"{entry.teacher_id},"
            f"{entry.class_id},"
            f"{entry.section_id or ''},"
            f"{entry.room_id or ''},"
            f"{entry.status},"
            f"{entry.remarks or ''}\n"
        )
    return output.getvalue()


def generate_teacher_timetable_pdf(entries: List[models.Timetable], teacher_name: str = "") -> bytes:
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, f"Teacher Timetable: {teacher_name}", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 8, f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}", ln=True, align="C")
    pdf.ln(5)

    pdf.set_font("Helvetica", "B", 8)
    col_widths = [18, 25, 30, 30, 30, 25, 25, 25, 25]
    headers = ["Day", "Period", "Start", "End", "Subject", "Class", "Section", "Room", "Status"]
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 7, h, border=1, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for entry in entries:
        row_data = [
            DAYS_OF_WEEK.get(entry.day_of_week, str(entry.day_of_week)),
            str(entry.period),
            entry.start_time or "",
            entry.end_time or "",
            str(entry.subject_id),
            str(entry.class_id),
            str(entry.section_id or "-"),
            str(entry.room_id or "-"),
            entry.status,
        ]
        for i, val in enumerate(row_data):
            pdf.cell(col_widths[i], 6, val, border=1, align="C")
        pdf.ln()

    return pdf.output(dest="S").encode("latin-1")


def generate_class_timetable_pdf(entries: List[models.Timetable], class_name: str = "") -> bytes:
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, f"Class Timetable: {class_name}", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 8, f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}", ln=True, align="C")
    pdf.ln(5)

    pdf.set_font("Helvetica", "B", 8)
    col_widths = [18, 25, 30, 30, 30, 25, 25, 25, 25]
    headers = ["Day", "Period", "Start", "End", "Subject", "Teacher", "Section", "Room", "Status"]
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 7, h, border=1, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for entry in entries:
        row_data = [
            DAYS_OF_WEEK.get(entry.day_of_week, str(entry.day_of_week)),
            str(entry.period),
            entry.start_time or "",
            entry.end_time or "",
            str(entry.subject_id),
            str(entry.teacher_id),
            str(entry.section_id or "-"),
            str(entry.room_id or "-"),
            entry.status,
        ]
        for i, val in enumerate(row_data):
            pdf.cell(col_widths[i], 6, val, border=1, align="C")
        pdf.ln()

    return pdf.output(dest="S").encode("latin-1")